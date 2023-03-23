
from pathlib import Path
import csv
import shutil
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles.borders import Border, Side
import pandas
import os
import datetime
import fnmatch

import user_interface.UI_master

# pip install pandas
# pip install openpyxl


FORMAT_TRANSAKSJONSOVERSIKT = 1
FORMAT_SOEK_I_TRANSAKSJONER = 2

template_path = 'template/Filterregnskap_template.xlsx'
header_row = 2  # The row where the category headers are
cell_with_year = 'C1'  # The cell where the current year is written
cell_with_name = 'B1'

class Transaciton:
    def __init__(self):
        self.date = ''
        self.bank_description = ''
        self.belop_inn = ''
        self.belop_ut = ''
        self.ref = ''
        self.num_ref = ''

    def equals(self, other):
        # If date and ref num are identical, we say that the transactions are the same
        date_is_equal = str(self.date) == str(other.date)
        ref_is_equal = str(self.ref) == str(other.ref)
        if len(str(self.num_ref)) == 0 or len(str(other.num_ref)) == 0 or other.num_ref is None:
            num_ref_is_equal = True
        else:
            num_ref_is_equal = str(self.num_ref) == str(other.num_ref)

        return date_is_equal and ref_is_equal and num_ref_is_equal

def insert_empty_rows(sheet, row_to_insert_at, num_rows_to_insert, date_col, description_col, attachment_col, nok_in_col, nok_out_col):

    if num_rows_to_insert == 0:
        return 0

    # When inserting blank rows, the cell merge properties are not shifted down. Loop over the affected rows and remember the merge properties
    cells_to_merge_after_insert = []
    for row in range(row_to_insert_at + 2, row_to_insert_at + 2 + num_rows_to_insert):
        merged_cells = list(sheet.merged_cells.ranges)
        for cell_range in merged_cells:
            bounds = cell_range.bounds
            if bounds[1] == row:
                sheet.merged_cells.remove(str(cell_range))
                cells_to_merge_after_insert.append(bounds)

    sheet.insert_rows(row_to_insert_at + 1, num_rows_to_insert)

    # Merge cells that were preciously merged
    for bounds in cells_to_merge_after_insert:
        sheet.merge_cells(  start_row=bounds[1] + num_rows_to_insert,
                            start_column=bounds[0],
                            end_row=bounds[3] + num_rows_to_insert,
                            end_column=bounds[2])

    # When inserting blank rows, the cell height is not shifted down. Loop over the affected rows and fix
    for row in range (row_to_insert_at + 10, row_to_insert_at, -1):  #
        sheet.row_dimensions[row + num_rows_to_insert].height = sheet.row_dimensions[
            row].height  # Transfer dimension changes of the shifted cells
        sheet.row_dimensions[row].height = None

    # When inserting blank rows, the cell merge properties are not shifted down. Loop over the affected rows and fix
    for row in range (row_to_insert_at + 1, row_to_insert_at + 1 + num_rows_to_insert):
        for cell_range in sheet.merged_cells.ranges:
            bounds = cell_range.bounds
            if bounds[1] == row:
                sheet.merged_cells.remove (str (cell_range))
                sheet.merge_cells (start_row=bounds[1] + num_rows_to_insert, start_column=bounds[0],
                                   end_row=bounds[3] + num_rows_to_insert, end_column=bounds[2])

    # Add cell borders for the new cells
    thin_border = Border (left=Side (style='thin'), right=Side (style='thin'), top=Side (style='thin'),
                          bottom=Side (style='thin'))
    for row in range (row_to_insert_at + 1, row_to_insert_at + 1 + num_rows_to_insert):
        sheet[date_col + str (row)].border = thin_border
        sheet[description_col + str (row)].border = thin_border
        sheet[attachment_col + str (row)].border = thin_border
        sheet[nok_in_col + str (row)].border = thin_border
        sheet[nok_out_col + str (row)].border = thin_border

    return num_rows_to_insert


def run_main_program(create_new_account, csv_transactions_file, year_to_track, account_name, account_filepath):
    global template_path
    global header_row
    global cell_with_year

    output_filepath = account_filepath
    if create_new_account:
        input_filepath = template_path
    else:
        input_filepath = account_filepath
    tmp_filepath = 'tmp/regnskap_tmp.xlsx'

    # Check for invalid csv file
    if not os.path.isfile(csv_transactions_file):
        success = False
        message = 'Ugyldig fil: ' + str(csv_transactions_file)
        title = 'Avbrutt'
        return success, message, title

    # Check for invalid input file location
    if not os.path.exists(os.path.dirname(input_filepath)):
        success = False
        message = 'Ugyldig mappe: ' + str(os.path.dirname(input_filepath))
        title = 'Avbrutt'
        return success, message, title
    shutil.copyfile(src=input_filepath, dst=tmp_filepath)

    # Check for invalid account name
    if create_new_account and len(account_name) == 0:
        success = False
        message = 'Oppgi navn på regnskap'
        title = 'Avbrutt'

    # Get format type of csv file
    format_type = -1
    fid = open (csv_transactions_file, 'r', encoding='cp1252')
    header_line = fid.readline(1000)
    fid.close ()
    if header_line.find('Bokføringsdato') != -1:
        format_type = FORMAT_TRANSAKSJONSOVERSIKT
    else:
        fid = open (csv_transactions_file, 'r', encoding='UTF-8')
        header_line = fid.readline(1000)
        fid.close()
        if header_line.find('Bokført') != -1:
            format_type = FORMAT_SOEK_I_TRANSAKSJONER
        else:
            success = False
            message = 'Ugyldig format på : ' + str(csv_transactions_file)
            title = 'Avbrutt'
            return success, message, title
    
    # Read csv-file exported from Sparebanken Sør
    if format_type == FORMAT_TRANSAKSJONSOVERSIKT:
        file = open(csv_transactions_file, 'r', encoding='cp1252')
    else:
        file = open (csv_transactions_file, 'r', encoding='UTF-8')
    csvreader = csv.reader(file, delimiter=';')

    # Read data from csv file
    csv_transactions_header = next(csvreader)
    csv_transactions = []
    if format_type == FORMAT_TRANSAKSJONSOVERSIKT:
        date_index = csv_transactions_header.index('Bokføringsdato')
        bank_description_index = csv_transactions_header.index('Beskrivelse')
        nok_in_index = csv_transactions_header.index('Inn på konto')
        nok_out_index = csv_transactions_header.index('Ut av konto')
        ref_index = csv_transactions_header.index('Ref.')
        num_ref_index = csv_transactions_header.index('Num.Ref.')

        for row in csvreader:
            if len(row) == 0 or row[date_index] == '':
                break

            transaction = Transaciton()
            transaction.date = row[date_index]
            transaction.bank_description = row[bank_description_index]
            transaction.belop_inn = row[nok_in_index]
            transaction.belop_ut = row[nok_out_index]
            transaction.ref = row[ref_index]
            transaction.num_ref = row[num_ref_index]
            csv_transactions.append(transaction)
    else:
        date_index = csv_transactions_header.index('Bokført')
        bank_description_index = csv_transactions_header.index('Beskrivelse')
        tekstkode_indeks = csv_transactions_header.index ('Tekstkode')
        belop_indexs = csv_transactions_header.index ('Beløp')
        ref_index = csv_transactions_header.index('Arkivref.')

        for row in csvreader:
            if len(row) == 0 or row[date_index] == '':
                break
            transaction = Transaciton()
            transaction.date = row[date_index]
            transaction.bank_description = row[tekstkode_indeks] + '  ' + row[bank_description_index]
            transaction.ref = row[ref_index]
            if len(row[belop_indexs]) > 0:
                if row[belop_indexs][0] == '-':
                    transaction.belop_ut = row[belop_indexs][1:]
                else:
                    transaction.belop_inn = row[belop_indexs]
            csv_transactions.append(transaction)
    file.close()

    # Read transactions in current accounting form
    workbook = openpyxl.load_workbook(filename=tmp_filepath)
    sheet = workbook['Regnskap']

    # Set account name and year
    if create_new_account:
        sheet[cell_with_name] = 'REGNSKAP ' + account_name.upper()
        sheet[cell_with_year] = year_to_track

    # If year is not listed in the account, return False
    year_to_track = sheet[cell_with_year].value
    if not str(year_to_track).isnumeric() or int(year_to_track) < 1900 or int(year_to_track) > 2500:
        success = False
        message = 'Error: Årstall (' + str(year_to_track) + ') er ugyldig'
        title = 'Avbrutt'
        workbook.close()
        return success, message, title


    # Read column indexes for the different categories
    date_col = ''
    description_col = ''
    attachment_col = ''
    nok_in_col = ''
    nok_out_col = ''
    category_col = ''
    category_ok_col = ''
    bank_description_col = ''
    ref_col = ''
    num_ref_col = ''
    all_cols_found = False
    for col in 'ABCDEFGHIJKLMNOPQRSTUVWXYZ':
        header_description = sheet[col + str(header_row)].value
        if header_description == 'Dato':
            date_col = col
        elif header_description == 'Beskrivelse':
            description_col = col
        elif header_description == 'Bilag':
            attachment_col = col
        elif header_description == 'INN':
            nok_in_col = col
        elif header_description == 'UT':
            nok_out_col = col
        elif header_description == 'Kategori':
            category_col = col
        elif header_description == 'Kategori OK':
            category_ok_col = col
        elif header_description == 'Beskrivelse fra Sparebanken Sør':
            bank_description_col = col
        elif header_description == 'Ref.':
            ref_col = col
        elif header_description == 'Num.Ref':
            num_ref_col = col

        if (date_col != '' and
                description_col != '' and
                attachment_col != '' and
                nok_in_col != '' and
                nok_out_col != '' and
                category_col != '' and
                category_ok_col != '' and
                bank_description_col != '' and
                ref_col != '' and
                num_ref_col != ''):
            all_cols_found = True
            break

    if not all_cols_found:
        success = False
        message = 'Ugyldig format på : ' + str(output_filepath) + '. En eller flere kolonner ble ikke funnet'
        title = 'Avbrutt'
        workbook.close()
        return success, message, title

    # Find row numbers for "IB Bank" and "UB Bank"
    all_relevant_cols = date_col + description_col + attachment_col + nok_in_col + nok_out_col + category_col + category_ok_col + bank_description_col + ref_col + num_ref_col
    UB_Bank_row = -1
    IB_Bank_row = -1
    row = header_row + 1
    while UB_Bank_row == -1:
        description = sheet[description_col + str(row)].value
        if description == 'IB Bank':
            IB_Bank_row = row
        if description == 'UB Bank':
            UB_Bank_row = row
            break
        row += 1
        if row >= 5000:
            break
    if IB_Bank_row == -1 or UB_Bank_row == -1:
        success = False
        message = 'Klarte ikke finne "IB Bank" og/eller "UB Bank" i kolonne ' + description_col + ' i eksisterende regnskap'
        title = 'Avbrutt'
        workbook.close()
        return success, message, title

    # Get row number of first and last recorded transaction
    last_transaction_row = -1
    row = UB_Bank_row
    while last_transaction_row == -1:
        row -= 1
        row_is_empty = True
        for col in all_relevant_cols:
            if sheet[col + str(row)].value is not None:
                row_is_empty = False
        if not row_is_empty:
            last_transaction_row = row
            break
    first_transaction_row = IB_Bank_row + 1


    # Read all old transactions from the account form
    old_transactions = []
    for row in range(first_transaction_row, last_transaction_row + 1):
        transaction = Transaciton()
        transaction.date = sheet[date_col + str(row)].value.strftime("%d.%m.%Y")
        transaction.bank_description = sheet[bank_description_col + str(row)].value
        transaction.belop_inn = sheet[nok_in_col + str(row)].value
        transaction.belop_ut = sheet[nok_out_col + str(row)].value
        transaction.ref = sheet[ref_col + str(row)].value
        transaction.num_ref = sheet[num_ref_col + str(row)].value
        old_transactions.append(transaction)



    # Iterate through csv transactions and pick out the ones from the correct year that are not already present in the old account form
    new_transactions = []
    for potentially_new_transaction in csv_transactions:
        # Check if transaction is from the correct year
        transaction_year = potentially_new_transaction.date.split('.')[2]
        if int(transaction_year) != int(year_to_track):
            continue

        # Check if transaction is old
        transaction_is_old = False
        for old_transaction in old_transactions:
            if potentially_new_transaction.equals(old_transaction):
                transaction_is_old = True
                break

        if not transaction_is_old:
            new_transactions.append(potentially_new_transaction)
    if len(new_transactions) == 0:
        success = False
        message = 'Alle transaksjoner i csv-dokumentet er allerede inkludert i regnskapet'
        title = 'Ingen transaksjoner funnet'
        return success, message, title

    # Insert blank rows to fill inn with the new transactions
    UB_Bank_row += insert_empty_rows(sheet, last_transaction_row, len(new_transactions), date_col, description_col, attachment_col, nok_in_col, nok_out_col)

    # Reverse new transactions, to get the oldest ones at the top
    new_transactions.reverse()

    # Write new transactions to account form
    first_new_transaction_row = last_transaction_row + 1
    row = first_new_transaction_row
    last_new_transaction_row = first_new_transaction_row - 1
    for transaction in new_transactions:
        # Write date
        date = transaction.date.split('.')
        date_cell = date_col + str(row)
        sheet[date_cell] = pandas.Timestamp(day=int(date[0]), month=int(date[1]), year=int(date[2]))
        sheet[date_cell].number_format = 'DD.MM.YYYY'

        # Write NOK in
        NOK_in = transaction.belop_inn
        if NOK_in != '':
            NOK_in = float(NOK_in.replace(',', '.'))
        nok_in_cell = nok_in_col + str(row)
        sheet[nok_in_cell] = NOK_in
        sheet[nok_in_cell].number_format = '[$kr-414]" "#,##0.00;[Red]"-"[$kr-414]" "#,##0.00'

        # Write NOK out
        NOK_out = transaction.belop_ut
        if NOK_out != '':
            NOK_out = float(NOK_out.replace(',', '.'))
        nok_out_cell = nok_out_col + str(row)
        sheet[nok_out_cell] = NOK_out
        sheet[nok_out_cell].number_format = '[$kr-414]" "#,##0.00;[Red]"-"[$kr-414]" "#,##0.00'

        # Set category to undefined
        category_cell = category_col + str(row)
        sheet[category_cell] = 'Udefinert'

        # Write 'Category OK' check
        category_ok_formula = '=AND(OR(AND((' + nok_in_cell + '>' + nok_out_cell + '), ISNUMBER(MATCH(' + category_cell + ', INN, 0))), AND((' + nok_in_cell + '<' + nok_out_cell + '), ISNUMBER(MATCH(' + category_cell + ', UT, 0)))), NOT(EXACT(REPLACE(' + category_cell + ', 1, 2, ""), "Udefinert")))'
        category_ok_cell = category_ok_col + str(row)
        sheet[category_ok_cell] = category_ok_formula

        # Write bank description
        sheet[bank_description_col + str(row)] = transaction.bank_description

        # Write Ref.
        sheet[ref_col + str(row)] = transaction.ref

        # Write Num.Ref
        sheet[num_ref_col + str(row)] = transaction.num_ref

        last_new_transaction_row = row
        row += 1


    if len(new_transactions) > 0:
        # Set conditional formatting in category_col
        redFill = openpyxl.styles.PatternFill(start_color='FD8787', end_color='FD8787', fill_type='solid')
        rule = FormulaRule(formula=['NOT(' + category_ok_col + str(first_transaction_row) + ')'], fill=redFill)
        # Delete old conditional formatting
        for key in list(sheet.conditional_formatting._cf_rules.keys()):
            rule_old = sheet.conditional_formatting._cf_rules.get(key)[0]
            if fnmatch.filter(rule_old.formula, rule.formula[0]):
                sheet.conditional_formatting._cf_rules.pop(key)
        cell_range = category_col + str(first_transaction_row) + ':' + category_col + str(last_new_transaction_row)
        sheet.conditional_formatting.add(cell_range, rule)

        # Add category drop-down menu in category_col
        dv = DataValidation(type='list', showDropDown=None, allowBlank=True, formula1='IF(' + nok_in_col + str(first_new_transaction_row) + '>' + nok_out_col + str(first_new_transaction_row) + ',INN,UT)')
        # Check if data validation rule already exists
        found_dv = False
        for dv_old in sheet.data_validations.dataValidation:
            if dv_old.type == dv.type and dv_old.showDropDown == dv.showDropDown and dv_old.allowBlank == dv.allowBlank:
                if fnmatch.filter([dv.formula1], 'IF(*>*,INN,UT)'):
                    dv = dv_old
                    found_dv = True
                    break
        if not found_dv:
            sheet.add_data_validation(dv)
        dv.add(category_col + str(first_new_transaction_row) + ':' + category_col + str(last_new_transaction_row))


    # Find row for BALANSESUM
    row = UB_Bank_row + 1
    BALANSESUM_row = -1
    while BALANSESUM_row == -1:
        if sheet[date_col + str(row)].value.find('BALANSESUM') != -1:
            BALANSESUM_row = row
        row += 1

    # Add formula for BALANSESUM
    sheet[nok_in_col + str(BALANSESUM_row)] = '=SUM(' + nok_in_col + '$' + str(IB_Bank_row) + ':' + nok_in_col + str(UB_Bank_row) + ')'
    sheet[nok_out_col + str(BALANSESUM_row)] = '=SUM(' + nok_out_col + '$' + str(IB_Bank_row) + ':' + nok_out_col + str(UB_Bank_row) + ')'
    sheet[category_col + str(BALANSESUM_row)] = '=' + nok_in_col + str(BALANSESUM_row) + '-' + nok_out_col + str(BALANSESUM_row)


    # Add 'Utgående Balanse' til sammendraget over regnskapet
    sheet_summary = workbook['Sammendrag']
    finished = False
    for col in range(1, 10):
        for row in range(1, 100):
            if sheet_summary.cell(row, col).value is not None and fnmatch.filter([sheet_summary.cell(row, col).value], '=CONCATENATE("Utgående balanse ", Regnskap!A*)'):
                sheet_summary.cell(row, col).value = '=CONCATENATE("Utgående balanse ", Regnskap!A' + str(UB_Bank_row) + ')'
                sheet_summary.cell(row, col+2).value = '=Regnskap!' + nok_out_col + str(UB_Bank_row)
                finished = True
                break
        if finished:
            break

    workbook.save(tmp_filepath)
    workbook.close()

    # Backup current form
    if not create_new_account:
        if not os.path.exists('backups'):
            os.makedirs('backups')
        [_, file_name] = os.path.split(input_filepath)
        backup_filename = datetime.datetime.now().strftime("%Y.%m.%d_kl_%H.%M.%S") + '_' + file_name
        shutil.copyfile(input_filepath, 'backups/' + backup_filename)
    shutil.copyfile(tmp_filepath, output_filepath)

    if not os.path.exists('tmp'):
        os.makedirs('tmp')

    # Save the default file name
    f = open('tmp/last_account_name.txt', 'w')
    f.write(account_name)
    f.close()

    # Save default file selection
    f = open('tmp/last_account_file.txt', 'w')
    f.write(account_filepath)
    f.close()

    success = True
    message = str(len(new_transactions)) + ' nye transaksjoner lagt til'
    title = 'Ferdig'
    return success, message, title


def main():
    downloads_dir = str(Path.home() / "Downloads")
    documents_dir = str(Path.home() / "Documents")
    
    default_create_new_account = True
    default_year = str(datetime.date.today().year)
    default_name = ''
    default_new_transactions_file_dir = downloads_dir
    default_account_location = documents_dir
    default_exsisting_form = ''
    
    # Get default exsisting account
    if os.path.isfile('tmp/last_account_file.txt'):
        f = open('tmp/last_account_file.txt')
        default_exsisting_form = f.read()
        default_account_location = os.path.dirname(default_exsisting_form)
        f.close()
        default_create_new_account = False

    # Get default account name
    if os.path.isfile('tmp/last_account_name.txt'):
        f = open('tmp/last_account_name.txt')
        default_name = f.read()
        f.close()

    user_interface.UI_master.run_GUI(default_create_new_account,
                                     default_year,
                                     default_name,
                                     default_new_transactions_file_dir,
                                     default_account_location,
                                     default_exsisting_form,
                                     run_main_program)

if __name__ == "__main__":
    main()
