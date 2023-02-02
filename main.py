
num_libraries = 16
library_num = 1
def print_loading_screen():
    global library_num
    global num_libraries
    end_char = '\r'
    if library_num == num_libraries:
        end_char = '\n'
    print('Laster inn ' + str(library_num) + '/' + str(num_libraries) + '...', end=end_char)
    library_num += 1


print_loading_screen()
import tkinter as tk
print_loading_screen()
from tkinter import filedialog
print_loading_screen()
from pathlib import Path
print_loading_screen()
import csv
print_loading_screen()
import shutil
print_loading_screen()
import openpyxl
print_loading_screen()
from openpyxl.styles import PatternFill
print_loading_screen()
from openpyxl.formatting.rule import FormulaRule
print_loading_screen()
from openpyxl.worksheet.datavalidation import DataValidation
print_loading_screen()
from openpyxl.styles.borders import Border, Side
print_loading_screen()
import pandas
print_loading_screen()
import os
print_loading_screen()
import datetime
print_loading_screen()
import pyautogui
print_loading_screen()
import easygui
print_loading_screen()
import fnmatch


# pip install easygui
# pip install pandas
# pip install openpyxl
# pip install pyautogui

header_row = 2  # The row where the category headers are
cell_with_year = 'C1'  # The cell where the current year is written
template_path = 'template/Filterregnskap_template.xlsx'


def get_year_to_track():
    year_to_track_local = easygui.enterbox(msg='Oppgi hvilket år som skal regnskapsføres: ', title='Velg år', default=str(datetime.date.today ().year))
    if (year_to_track_local is None) or (not year_to_track_local.isnumeric()):
        raise Exception('Ugyldig format på oppgitt årstall. Avslutter uten å endre regnearket.')
    return year_to_track_local

year_to_track = None
create_new_account = False

downloads_dir = str(Path.home() / "Downloads")
documents_dir = str(Path.home() / "Documents")

transactions_file_default_dir = downloads_dir
existing_form_default_dir = documents_dir
existing_form_default_name = ''

if not os.path.exists('tmp'):
    os.makedirs('tmp')

# Get default file selection
if os.path.isfile('tmp/last_account_file.txt'):
    f = open('tmp/last_account_file.txt')
    existing_form_default_filepath = f.read()
    [existing_form_default_dir, existing_form_default_name] = os.path.split(existing_form_default_filepath)
    f.close()

root = tk.Tk()
root.withdraw()
# Get filepath to csv exported from Sparebanken Sor
transactions_file_path = filedialog.askopenfilename(title='Velg transaksjonsoversikt eksportert fra Sparebanken Sør', initialdir=transactions_file_default_dir)
# If no file was selected, abort
if transactions_file_path is None or transactions_file_path == '':
    pyautogui.alert('Ingen fil valgt, avbryter scriptet', 'Avbrutt')
    exit()

existing_form = filedialog.askopenfilename(title='Velg regnskap å oppdatere', initialdir=existing_form_default_dir, initialfile=existing_form_default_name)

# If no file was selected, create a new one from template
if existing_form is None or existing_form == '':
    create_new_account = True
    year_to_track = get_year_to_track()
    new_account_location = filedialog.asksaveasfilename(title='Lagre nytt regnskap', initialdir=documents_dir, initialfile='Filterregnskap_' + str(year_to_track) + '.xlsx')
    if new_account_location is None or new_account_location == '':
        pyautogui.alert('Ugyldig fillokasjon, avbryter scriptet', 'Avbrutt')
        exit()

    shutil.copyfile(template_path, new_account_location)
    existing_form = new_account_location


# Save default file selection
f = open('tmp/last_account_file.txt', 'w')
f.write(existing_form)
f.close()

# Backup current form
if not create_new_account:
    if not os.path.exists('backups'):
        os.makedirs('backups')
    [_, file_name] = os.path.split(existing_form)
    backup_filename = datetime.datetime.now().strftime("%Y.%m.%d_kl_%H.%M.%S") + '_' + file_name
    shutil.copyfile(existing_form, 'backups/' + backup_filename)

# Make changes to the original document
output_file_path = existing_form

# Get format type of csv
FORMAT_TRANSAKSJONSOVERSIKT = 1
FORMAT_SOEK_I_TRANSAKSJONER = 2
def get_csv_format_type(transactions_file):
    fid = open (transactions_file_path, 'r', encoding='cp1252')
    header_line = fid.readline(1000)
    fid.close ()
    if header_line.find('Bokføringsdato') != -1:
        return FORMAT_TRANSAKSJONSOVERSIKT

    fid = open (transactions_file_path, 'r', encoding='UTF-8')
    header_line = fid.readline(1000)
    fid.close()
    if header_line.find('Bokført') != -1:
        return FORMAT_SOEK_I_TRANSAKSJONER
    raise Exception('Invalid format on ' + transactions_file)

format_type = get_csv_format_type(transactions_file_path)

# Read csv-file exported from Sparebanken Sør
if format_type == FORMAT_TRANSAKSJONSOVERSIKT:
    file = open(transactions_file_path, 'r', encoding='cp1252')
elif format_type == FORMAT_SOEK_I_TRANSAKSJONER:
    file = open (transactions_file_path, 'r', encoding='UTF-8')
csvreader = csv.reader(file, delimiter=';')
csv_transactions_header = next(csvreader)


if format_type == FORMAT_TRANSAKSJONSOVERSIKT:
    date_index = csv_transactions_header.index('Bokføringsdato')
    bank_description_index = csv_transactions_header.index('Beskrivelse')
    nok_in_index = csv_transactions_header.index('Inn på konto')
    nok_out_index = csv_transactions_header.index('Ut av konto')
    ref_index = csv_transactions_header.index('Ref.')
    num_ref_index = csv_transactions_header.index('Num.Ref.')
elif format_type == FORMAT_SOEK_I_TRANSAKSJONER:
    date_index = csv_transactions_header.index('Bokført')
    bank_description_index = csv_transactions_header.index('Beskrivelse')
    tekstkode_indeks = csv_transactions_header.index ('Tekstkode')
    belop_indexs = csv_transactions_header.index ('Beløp')
    ref_index = csv_transactions_header.index('Arkivref.')
    nok_in_index = len(csv_transactions_header)
    nok_out_index = len(csv_transactions_header) + 1
    num_ref_index = len(csv_transactions_header) + 2
    csv_transactions_header.append('Inn på konto')
    csv_transactions_header.append('Ut av konto')
    csv_transactions_header.append('Num.Ref.')


csv_transactions = []
for row in csvreader:
    if len(row) == 0 or row[date_index] == '':
        break
    else:
        if format_type == FORMAT_SOEK_I_TRANSAKSJONER:
            row[bank_description_index] = row[tekstkode_indeks] + '  ' + row[bank_description_index]
            row.append('') # Inn på konto
            row.append('') # Ut av konto
            row.append('') # Num.Ref
            if len(row[belop_indexs]) > 0:
                if row[belop_indexs][0] == '-':
                    row[nok_out_index] = row[belop_indexs][1:]
                else:
                    row[nok_in_index] = row[belop_indexs]

        csv_transactions.append(row)
file.close()



# Read current accounting form
workbook = openpyxl.load_workbook(filename=output_file_path)
sheet = workbook['Regnskap']

date_col = ''
description_col = ''
attachment_co = ''
nok_in_col = ''
nok_out_col = ''
category_col = ''
category_ok_col = ''
bank_description_col = ''
ref_col = ''
num_ref_col = ''


cols_alphabeth = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
all_cols_found = False
i = 0
while not all_cols_found:
    if i >= len(cols_alphabeth):
        raise Exception('Failed to parse existing account form')

    current_col = cols_alphabeth[i]
    header_description = sheet[current_col + str(header_row)].value
    if header_description == 'Dato':
        date_col = current_col
    elif header_description == 'Beskrivelse':
        description_col = current_col
    elif header_description == 'Bilag':
        attachment_co = current_col
    elif header_description == 'INN':
        nok_in_col = current_col
    elif header_description == 'UT':
        nok_out_col = current_col
    elif header_description == 'Kategori':
        category_col = current_col
    elif header_description == 'Kategori OK':
        category_ok_col = current_col
    elif header_description == 'Beskrivelse fra Sparebanken Sør':
        bank_description_col = current_col
    elif header_description == 'Ref.':
        ref_col = current_col
    elif header_description == 'Num.Ref':
        num_ref_col = current_col

    if (date_col != '' and
            description_col != '' and
            attachment_co != '' and
            nok_in_col != '' and
            nok_out_col != '' and
            category_col != '' and
            category_ok_col != '' and
            bank_description_col != '' and
            ref_col != '' and
            num_ref_col != ''):
        all_cols_found = True

    i += 1


all_relevant_cols = date_col + description_col + attachment_co + nok_in_col + nok_out_col + category_col + category_ok_col + bank_description_col + ref_col + num_ref_col
UB_Bank_row = -1
IB_Bank_row = -1
row = header_row + 1
while UB_Bank_row == -1:
    description = sheet[description_col + str(row)].value
    if description == 'IB Bank':
        IB_Bank_row = row
    if description == 'UB Bank':
        UB_Bank_row = row
        break
    row += 1
    if row >= 5000:
        break

if IB_Bank_row == -1 or UB_Bank_row == -1:
    raise Exception('Failed to find "IB Bank" and/or "UB Bank" in column' + description_col + 'of existing account form')

# Get row number of last recorded transaction
last_transaction_row = -1
row = UB_Bank_row
while last_transaction_row == -1:
    row -= 1
    row_is_empty = True
    for col in all_relevant_cols:
        if sheet[col + str(row)].value is not None:
            row_is_empty = False
    if not row_is_empty:
        last_transaction_row = row
        break

# Get row number of first recorded transaction
first_transaction_row = IB_Bank_row + 1


# Read all old transactions from the account form, and store on the same format as the csv transactions
old_transactions = []
for row in range(first_transaction_row, last_transaction_row + 1):
    transaction = [0] * len(csv_transactions_header)
    transaction[date_index] = sheet[date_col + str(row)].value.strftime("%d.%m.%Y")
    transaction[bank_description_index] = sheet[bank_description_col + str(row)].value
    transaction[nok_in_index] = sheet[nok_in_col + str(row)].value
    transaction[nok_out_index] = sheet[nok_out_col + str(row)].value
    transaction[ref_index] = sheet[ref_col + str(row)].value
    transaction[num_ref_index] = sheet[num_ref_col + str(row)].value
    old_transactions.append(transaction)

# Get year to make account for
if year_to_track is None:
    year_to_track = sheet[cell_with_year].value
# If year is not listed in the account, ask the user what year to use
if not str(year_to_track).isnumeric() or int(year_to_track) < 1900 or int(year_to_track) > 2500:
    year_to_track = get_year_to_track()
sheet[cell_with_year] = year_to_track


# Iterate through csv transactions and pick out the ones from the correct year that are not already present in the old account form
new_transactions = []
for potentially_new_transaction in csv_transactions:
    # Check if transaction is from the correct year
    transaction_year = potentially_new_transaction[date_index].split('.')[2]
    if int(transaction_year) != int(year_to_track):
        continue

    # Check if transaction is old
    transaction_is_old = False
    for old_transaction in old_transactions:
        # If date and ref num are identical, we say that the transactions are the same
        date_differs = str(potentially_new_transaction[date_index]) != str(old_transaction[date_index])
        ref_differs = str(potentially_new_transaction[ref_index]) != str(old_transaction[ref_index])
        if len(str(potentially_new_transaction[num_ref_index])) == 0\
                or len(str(old_transaction[num_ref_index])) == 0\
                or old_transaction[num_ref_index] is None:
            num_ref_differs = False
        else:
            num_ref_differs = str(potentially_new_transaction[num_ref_index]) != str(old_transaction[num_ref_index])

        if not (date_differs or ref_differs or num_ref_differs):
            transaction_is_old = True
            break
    if not transaction_is_old:
        new_transactions.append(potentially_new_transaction)



# Insert blank rows to fill inn with the new transactions
if len(new_transactions) != 0:
    # When inserting blank rows, the cell merge properties are not shifted down. Loop over the affected rows and remember the merge properties
    cells_to_merge_after_insert = []
    for row in range(last_transaction_row + 2, last_transaction_row + 2 + len(new_transactions)):
        for cell_range in sheet.merged_cells.ranges:
            bounds = cell_range.bounds
            if bounds[1] == row:
                sheet.merged_cells.remove(str(cell_range))
                cells_to_merge_after_insert.append(bounds)

    sheet.insert_rows(last_transaction_row + 1, len(new_transactions))
    UB_Bank_row += len(new_transactions)

    # Merge cells that were preciously merged
    for bounds in cells_to_merge_after_insert:
        sheet.merge_cells (start_row=bounds[1] + len(new_transactions),
            start_column=bounds[0],
            end_row=bounds[3] + len(new_transactions),
            end_column=bounds[2])


    # When inserting blank rows, the cell height is not shifted down. Loop over the affected rows and fix
    for row in range(last_transaction_row + 10, last_transaction_row, -1): #
        sheet.row_dimensions[row + len(new_transactions)].height = sheet.row_dimensions[row].height # Transfer dimension changes of the shifted cells
        sheet.row_dimensions[row].height = None

    # When inserting blank rows, the cell merge properties are not shifted down. Loop over the affected rows and fix
    for row in range (last_transaction_row + 1, last_transaction_row + 1 + len (new_transactions)):
        for cell_range in sheet.merged_cells.ranges:
            bounds = cell_range.bounds
            if bounds[1] == row:
                sheet.merged_cells.remove(str(cell_range))
                sheet.merge_cells(start_row=bounds[1] + len(new_transactions), start_column=bounds[0], end_row=bounds[3] + len(new_transactions), end_column=bounds[2])

    # Add cell borders for the new cells
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for row in range(last_transaction_row + 1, last_transaction_row + 1 + len(new_transactions)):
        sheet[date_col + str(row)].border = thin_border
        sheet[description_col + str(row)].border = thin_border
        sheet[attachment_co + str(row)].border = thin_border
        sheet[nok_in_col + str(row)].border = thin_border
        sheet[nok_out_col + str(row)].border = thin_border

# Reverse new transactions, to get the oldest ones at the top
new_transactions.reverse()

# Write new transactions to account form
first_new_transaction_row = last_transaction_row + 1
row = first_new_transaction_row
last_new_transaction_row = first_new_transaction_row - 1
for transaction in new_transactions:
    # Write date
    date = transaction[date_index].split('.')
    date_cell = date_col + str(row)
    sheet[date_cell] = pandas.Timestamp(day=int(date[0]), month=int(date[1]), year=int(date[2]))
    sheet[date_cell].number_format = 'DD.MM.YYYY'

    # Write NOK in
    NOK_in = transaction[nok_in_index]
    if NOK_in != '':
        NOK_in = float(NOK_in.replace(',', '.'))
    nok_in_cell = nok_in_col + str(row)
    sheet[nok_in_cell] = NOK_in
    sheet[nok_in_cell].number_format = '[$kr-414]" "#,##0.00;[Red]"-"[$kr-414]" "#,##0.00'

    # Write NOK out
    NOK_out = transaction[nok_out_index]
    if NOK_out != '':
        NOK_out = float(NOK_out.replace(',', '.'))
    nok_out_cell = nok_out_col + str(row)
    sheet[nok_out_cell] = NOK_out
    sheet[nok_out_cell].number_format = '[$kr-414]" "#,##0.00;[Red]"-"[$kr-414]" "#,##0.00'

    # Set category to undefined
    category_cell = category_col + str(row)
    sheet[category_cell] = 'Udefinert'

    # Write 'Category OK' check
    category_ok_formula = '=AND(OR(AND((' + nok_in_cell + '>' + nok_out_cell + '), EXACT(LEFT(' + category_cell + '), "+")), AND((' + nok_in_cell + '<' + nok_out_cell + '), EXACT(LEFT(' + category_cell + '), "÷"))), NOT(EXACT(REPLACE(' + category_cell + ', 1, 2, ""), "Udefinert")))'
    category_ok_cell = category_ok_col + str(row)
    sheet[category_ok_cell] = category_ok_formula

    # Write bank description
    sheet[bank_description_col + str(row)] = transaction[bank_description_index]

    # Write Ref.
    sheet[ref_col + str(row)] = transaction[ref_index]

    # Write Num.Ref
    sheet[num_ref_col + str(row)] = transaction[num_ref_index]

    last_new_transaction_row = row
    row += 1

if len(new_transactions) > 0:

    # Set conditional formatting in category_col
    redFill = openpyxl.styles.PatternFill(start_color='FD8787', end_color='FD8787', fill_type='solid')
    rule = FormulaRule(formula=['NOT(' + category_ok_col + str(first_transaction_row) + ')'], fill=redFill)
    # Delete old conditional formatting
    for key in list(sheet.conditional_formatting._cf_rules.keys()):
        rule_old = sheet.conditional_formatting._cf_rules.get(key)[0]
        if fnmatch.filter(rule_old.formula, rule.formula[0]):
            sheet.conditional_formatting._cf_rules.pop(key)
    cell_range = category_col + str(first_transaction_row) + ':' + category_col + str(last_new_transaction_row)
    sheet.conditional_formatting.add(cell_range, rule)

    # Add category drop-down menu in category_col
    dv = DataValidation(type='list', showDropDown=None, allowBlank=True, formula1='IF(' + nok_in_col + str(first_new_transaction_row) + '>' + nok_out_col + str(first_new_transaction_row) + ',INN,UT)')
    # Check if data validation rule already exists
    found_dv = False
    for dv_old in sheet.data_validations.dataValidation:
        if dv_old.type == dv.type and dv_old.showDropDown == dv.showDropDown and dv_old.allowBlank == dv.allowBlank:
            if fnmatch.filter([dv.formula1], 'IF(*>*,INN,UT)'):
                dv = dv_old
                found_dv = True
                break
    if not found_dv:
        sheet.add_data_validation(dv)
    dv.add(category_col + str(first_new_transaction_row) + ':' + category_col + str(last_new_transaction_row))


# Find row for BALANSESUM
row = UB_Bank_row + 1
BALANSESUM_row = -1
while BALANSESUM_row == -1:
    if sheet[date_col + str(row)].value.find('BALANSESUM') != -1:
        BALANSESUM_row = row
    row += 1

# Add formula for BALANSESUM
sheet[nok_in_col + str(BALANSESUM_row)] = '=SUM(' + nok_in_col + '$' + str(IB_Bank_row) + ':' + nok_in_col + str(UB_Bank_row) + ')'
sheet[nok_out_col + str(BALANSESUM_row)] = '=SUM(' + nok_out_col + '$' + str(IB_Bank_row) + ':' + nok_out_col + str(UB_Bank_row) + ')'
sheet[category_col + str(BALANSESUM_row)] = '=' + nok_in_col + str(BALANSESUM_row) + '-' + nok_out_col + str(BALANSESUM_row)


# Add 'Utgående Balanse' til sammendraget over regnskapet
sheet_summary = workbook['Sammendrag']
finished = False
for col in range(1, 10):
    for row in range(1, 100):
        if sheet_summary.cell(row, col).value is not None and fnmatch.filter([sheet_summary.cell(row, col).value], '=CONCATENATE("Utgående balanse ", Regnskap!A*)'):
            sheet_summary.cell(row, col).value = '=CONCATENATE("Utgående balanse ", Regnskap!A' + str(UB_Bank_row) + ')'
            sheet_summary.cell(row, col+2).value = '=Regnskap!' + nok_out_col + str(UB_Bank_row)
            finished = True
            break
    if finished:
        break


workbook.save(output_file_path)
workbook.close()


pyautogui.alert(str(len(new_transactions)) + ' new transactions added', 'Script finished successfully')

